"""
CodeMao CRM 一键抓取脚本
输入: TSV 文件 (user_id\tchild_name\tcourse_name)
输出: 格式化 Excel

用法: python quick_excel.py <tsv_file> <output_xlsx>
或作为模块导入: from quick_excel import tsv_to_excel
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def tsv_to_excel(tsv_text, output_path):
    """将 TSV 文本直接转为格式化 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "未完课数据"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill('solid', fgColor='4472C4')
    ha = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(["用户ID", "孩子姓名", "课程名称"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, border

    rows = [line.split('\t') for line in tsv_text.strip().split('\n') if line.strip()]
    for ri, vals in enumerate(rows, 2):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=int(v) if ci == 1 and v.isdigit() else v)
            c.border = border
            c.alignment = Alignment(vertical='center')

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    ws.auto_filter.ref = f"A1:C{len(rows)+1}"

    wb.save(output_path)
    return len(rows)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        n = tsv_to_excel(f.read(), sys.argv[2])
    print(f"Saved: {sys.argv[2]} ({n} records)")
